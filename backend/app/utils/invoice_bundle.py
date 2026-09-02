"""
SIAE — Generador de Reporte en Excel y Bundle ZIP de Facturas.
Construye en memoria (RAM) el libro de cálculo con fórmulas dinámicas y empaqueta
los archivos digitales (XML, PDF y Devoluciones) con índice correlativo para auditoría.
"""

import io
import os
import zipfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _get_existing_path(path_or_url: str | None, default_dir: str = "uploads") -> str | None:
    """Resuelve la ruta física del archivo en disco soportando rutas relativas y absolutas."""
    if not path_or_url:
        return None
    clean = path_or_url.lstrip("/")
    if os.path.exists(clean):
        return clean
    if os.path.exists(os.path.join("/app", clean)):
        return os.path.join("/app", clean)
    direct = os.path.join(default_dir, os.path.basename(path_or_url))
    if os.path.exists(direct):
        return direct
    if os.path.exists(os.path.join("/app", direct)):
        return os.path.join("/app", direct)
    return None


def build_comprobacion_excel(
    folio_tramite: str,
    facturas_sorted: list,
    tramite_type: str = "viatico"
) -> bytes:
    """
    Construye en memoria el libro de Excel (.xlsx) con metadatos fiscales,
    formato profesional y fórmulas nativas de cálculo (SUM y SUMIF).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Comprobación"
    ws.views.sheetView[0].showGridLines = True

    # ── Paleta de Estilos y Colores ──
    HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Azul Institucional
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    SUBHEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    
    TITLE_FONT = Font(name="Calibri", size=13, bold=True, color="1E3A8A")
    META_FONT = Font(name="Calibri", size=10, bold=False, color="4B5563")
    META_BOLD = Font(name="Calibri", size=10, bold=True, color="1F2937")
    
    DATA_FONT = Font(name="Calibri", size=10, color="111827")
    ZEBRA_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    TOTAL_FILL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    TOTAL_FONT = Font(name="Calibri", size=11, bold=True, color="111827")
    
    CAT_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    CAT_FONT = Font(name="Calibri", size=10, bold=True, color="374151")

    THIN_BORDER_SIDE = Side(border_style="thin", color="D1D5DB")
    CELL_BORDER = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)
    
    DOUBLE_BOTTOM_SIDE = Side(border_style="double", color="111827")
    TOP_THIN_SIDE = Side(border_style="thin", color="111827")
    TOTAL_BORDER = Border(top=TOP_THIN_SIDE, bottom=DOUBLE_BOTTOM_SIDE, left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE)

    # ── 1. Encabezado y Metadatos ──
    tipo_nombre = "Comisión de Viáticos" if tramite_type == "viatico" else "Gasto a Reserva de Comprobar (GRC)"
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"SIAE — REPORTE DE COMPROBACIÓN DE FACTURAS ({tipo_nombre.upper()})"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Folio del Trámite:"
    ws["A2"].font = META_BOLD
    ws["B2"] = str(folio_tramite)
    ws["B2"].font = META_FONT
    
    ws["D2"] = "Fecha de Generación:"
    ws["D2"].font = META_BOLD
    ws["E2"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["E2"].font = META_FONT

    ws["G2"] = "Total de Comprobantes:"
    ws["G2"].font = META_BOLD
    ws["H2"] = len(facturas_sorted)
    ws["H2"].font = META_FONT
    ws["H2"].alignment = Alignment(horizontal="left")

    ws.row_dimensions[3].height = 10

    # ── 2. Columnas de la Tabla ──
    headers = [
        "#",
        "Fecha de Emisión",
        "Folio del Trámite",
        "Proveedor / Emisor",
        "RFC Emisor",
        "Folio Fiscal (UUID)",
        "Serie y Folio Interno",
        "Monto Total",
        "Estado SAT",
        "Categoría de Gasto",
        "Descripción / Concepto"
    ]

    header_row_idx = 4
    ws.row_dimensions[header_row_idx].height = 24
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=header_text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center" if col_idx in [1, 2, 3, 5, 7, 9] else "left", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER

    # ── 3. Llenado de Filas de Facturas ──
    start_data_row = 5
    current_row = start_data_row
    categories_used = []

    for idx, f in enumerate(facturas_sorted, start=1):
        # Fecha
        if f.fecha_emision:
            fecha_str = f.fecha_emision.strftime("%Y-%m-%d")
        else:
            fecha_str = "N/A"

        # Serie y Folio
        serie_folio = f"{f.serie or ''} {f.folio or ''}".strip() or "N/A"
        
        # Categoría
        cat_name = f.category.name if (f.category and hasattr(f.category, "name")) else "Sin Categoría"
        if cat_name not in categories_used:
            categories_used.append(cat_name)

        row_values = [
            idx,
            fecha_str,
            str(folio_tramite),
            f.emisor_nombre or "",
            f.emisor_rfc or "",
            f.uuid or "CARGA MANUAL",
            serie_folio,
            float(f.total or 0.0),
            f.sat_status or "Desconocido",
            cat_name,
            f.description or ""
        ]

        ws.row_dimensions[current_row].height = 20
        is_zebra = (idx % 2 == 0)

        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = CELL_BORDER
            if is_zebra:
                cell.fill = ZEBRA_FILL
            
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [2, 3, 5, 7, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 8:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        current_row += 1

    last_data_row = current_row - 1

    # ── 4. Totales y Fórmulas Dinámicas ──
    if facturas_sorted:
        current_row += 1
        
        # Fila de Total Comprobado
        ws.row_dimensions[current_row].height = 22
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        label_cell = ws.cell(row=current_row, column=1, value="TOTAL COMPROBADO:")
        label_cell.font = TOTAL_FONT
        label_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        for c in range(1, 8):
            ws.cell(row=current_row, column=c).fill = TOTAL_FILL
            ws.cell(row=current_row, column=c).border = TOTAL_BORDER

        total_formula_cell = ws.cell(row=current_row, column=8)
        total_formula_cell.value = f"=SUM(H{start_data_row}:H{last_data_row})"
        total_formula_cell.font = TOTAL_FONT
        total_formula_cell.number_format = '"$"#,##0.00'
        total_formula_cell.alignment = Alignment(horizontal="right", vertical="center")
        total_formula_cell.fill = TOTAL_FILL
        total_formula_cell.border = TOTAL_BORDER

        for c in range(9, 12):
            ws.cell(row=current_row, column=c).fill = TOTAL_FILL
            ws.cell(row=current_row, column=c).border = TOTAL_BORDER

        current_row += 2

        # Desglose por Categorías con SUMIF
        ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=8)
        cat_header = ws.cell(row=current_row, column=5, value="DESGLOSE POR CATEGORÍA DE GASTO")
        cat_header.font = SUBHEADER_FONT
        cat_header.fill = SUBHEADER_FILL
        cat_header.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        for cat in categories_used:
            ws.row_dimensions[current_row].height = 19
            ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
            c_label = ws.cell(row=current_row, column=5, value=cat)
            c_label.font = CAT_FONT
            c_label.alignment = Alignment(horizontal="left", vertical="center")
            
            for c in range(5, 8):
                ws.cell(row=current_row, column=c).fill = CAT_FILL
                ws.cell(row=current_row, column=c).border = CELL_BORDER

            c_val = ws.cell(row=current_row, column=8)
            c_val.value = f'=SUMIF(J{start_data_row}:J{last_data_row}, "{cat}", H{start_data_row}:H{last_data_row})'
            c_val.font = CAT_FONT
            c_val.number_format = '"$"#,##0.00'
            c_val.alignment = Alignment(horizontal="right", vertical="center")
            c_val.fill = CAT_FILL
            c_val.border = CELL_BORDER
            current_row += 1

    # ── 5. Ajuste Dinámico de Anchos de Columna ──
    col_widths = {
        "A": 6,   # #
        "B": 16,  # Fecha
        "C": 18,  # Folio
        "D": 35,  # Emisor
        "E": 16,  # RFC
        "F": 38,  # UUID
        "G": 20,  # Serie/Folio
        "H": 18,  # Total
        "I": 15,  # SAT
        "J": 24,  # Categoría
        "K": 30   # Concepto
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


def create_invoices_zip_bundle(
    folio: str,
    facturas: list,
    tramite_type: str = "viatico",
    extra_files: list[tuple[str, str]] | None = None
) -> io.BytesIO:
    """
    Construye en memoria el paquete ZIP contable:
    - Incluye el reporte en Excel en la raíz: reporte_{folio}.xlsx
    - Renombra facturas vinculadas con prefijo correlativo 1 a 1: 01_UUID.pdf, 01_UUID.xml...
    - Incluye comprobantes extras o devoluciones en subcarpeta Devoluciones/ o Extras/
    """
    # 1. Orden cronológico por fecha de emisión ascendente
    facturas_sorted = sorted(
        facturas,
        key=lambda x: (
            x.fecha_emision.replace(tzinfo=None) if (x.fecha_emision and hasattr(x.fecha_emision, "replace")) else (x.fecha_emision or datetime(1970, 1, 1)),
            x.id
        )
    )

    # 2. Generar Excel en memoria
    excel_bytes = build_comprobacion_excel(folio, facturas_sorted, tramite_type=tramite_type)

    # 3. Empaquetar en contenedor ZIP
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        # A. Guardar Excel en la raíz
        excel_filename = f"reporte_{folio}.xlsx"
        zip_file.writestr(excel_filename, excel_bytes)

        # B. Empaquetar facturas correlacionadas con prefijo (01_..., 02_...)
        default_folder = "viaticos" if tramite_type == "viatico" else "gastos_reserva_comprobar"
        for idx, invoice in enumerate(facturas_sorted, start=1):
            prefix = f"{idx:02d}"

            # XML
            if invoice.xml_filename:
                xml_path = _get_existing_path(invoice.xml_filename, f"uploads/{default_folder}/xml")
                if xml_path and os.path.exists(xml_path):
                    base_xml = os.path.basename(xml_path)
                    arc_xml = f"{prefix}_{base_xml}"
                    zip_file.write(xml_path, arcname=arc_xml)

            # PDF
            if invoice.pdf_filename:
                pdf_path = _get_existing_path(invoice.pdf_filename, f"uploads/{default_folder}/pdf")
                if pdf_path and os.path.exists(pdf_path):
                    base_pdf = os.path.basename(pdf_path)
                    arc_pdf = f"{prefix}_{base_pdf}"
                    zip_file.write(pdf_path, arcname=arc_pdf)

        # C. Anexos / Comprobantes de Devolución
        if extra_files:
            for file_path, arc_subpath in extra_files:
                if file_path:
                    valid_p = _get_existing_path(file_path, "uploads")
                    if valid_p and os.path.exists(valid_p):
                        zip_file.write(valid_p, arcname=arc_subpath)

    zip_buffer.seek(0)
    return zip_buffer
